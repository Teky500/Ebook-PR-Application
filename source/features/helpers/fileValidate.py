import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import os
import yaml
import pandas as pd


class FileTemplate:
    
    def __init__(self, f_path):
        self.file_path = f_path

        self.FILE_SHEET = "PA-Rights"
        with open('source/config/config.yaml', 'r') as config_file:
            yaml_file = yaml.safe_load(config_file)
            uni = yaml_file['University'] 


        self.FINAL_HEADER_FIELDS = {
                            "A3": "Title",
                            "B3": "Publisher",
                            "C3": "Platform_YOP", #int
                            "D3": "Platform_eISBN", #int
                            "E3": "OCN",            #int
                            "F3": "agreement_code",
                            "G3": "collection_name",
                            "H3": "title_metadata_last_modified", #date
                            "I3": uni
                        }
        
        
        self.NECESSARY_FIELDS = {
                        "A1": "Platform"
                        }

    def get_file_path(self):
        return self.file_path
    
    def get_file_sheet(self):
        return self.FILE_SHEET
    
    def get_necessary_fields(self):
        return self.NECESSARY_FIELDS
    
    def get_final_header_fields(self):
        return self.FINAL_HEADER_FIELDS
    
    def number_of_fixed_fields(self):
        return len(self.FINAL_HEADER_FIELDS)
    
    def set_file_path(self, f_path):
        self.file_path = f_path
    
    def set_file_sheet(self, f_sheet):
        self.FILE_SHEET = f_sheet
    
    def set_necessary_fields(self, fields):
        self.NECESSARY_FIELDS = fields
    
    def set_final_header_fields(self, fields):
        self.FINAL_HEADER_FIELDS = fields


class FileValidator:
    f = None
    error_message = []
    wb = None
    ws = None

    def __init__(self, file_to_be_validated):
        self.f = file_to_be_validated
        self.error_message = []

    def file_can_be_accessed(self):
    
        path = self.f.get_file_path()
        sheet = self.f.get_file_sheet()

        file_name, file_extension = os.path.splitext(path)
        if (path == ""):
            self.error_message.append("Error: No File Provided")

        elif file_extension not in [".xlsx", ".xls"]:
            self.error_message.append("Error: Invalid File")
        elif (not os.path.isfile(path)):
            self.error_message.append("Error: File Does Not Exist")
        else:
            try:
                self.wb = openpyxl.load_workbook(path)
                self.ws = self.wb[sheet]

            except (KeyError):
                self.error_message.append("Invalid Sheet name: Set sheet name to PA-Rights")
            
        return (len(self.error_message) == 0)
        

    def necessary_fields_not_empty(self):
        
        list2 = []

        for i in range (len(self.f.get_necessary_fields())):

            cell_keys = list(self.f.get_necessary_fields().keys())
            cell_values = list(self.f.get_necessary_fields().values())


            current_cell_value = self.ws[cell_keys[i]].value
            maybe_missing = cell_values[i]

            if (current_cell_value == None):
                list2.append("Missing Field: Insert " + maybe_missing + " field into cell " + cell_keys[i])

        self.error_message = self.error_message + list2


    def correct_header_format(self):
        
        list2 = []

        for i in range (len(self.f.get_final_header_fields())):

            cell_keys = list(self.f.get_final_header_fields().keys())
            cell_values = list(self.f.get_final_header_fields().values())


            current_cell_value = self.ws[cell_keys[i]].value
            expected_cell_value = cell_values[i]

            if (current_cell_value != expected_cell_value):
                list2.append("Invalid cell field: Set cell " + cell_keys[i] + " to " + cell_values[i])

        self.error_message = self.error_message + list2


    def get_error_message(self):
        return self.error_message
    

    def is_valid_file(self):

        if (self.file_can_be_accessed()):
            self.necessary_fields_not_empty()
            self.correct_header_format()

        return (len(self.error_message) == 0)


good_file1 = "exampleExcelFiles/UPEI_Ebooks_local_correct_sample1.xlsx"
good_file2 = "exampleExcelFiles/UPEI_Ebooks_local_correct_sample2.xlsx"
bad_file1 = "exampleExcelFiles/UPEI_Ebooks_local_incorrect_sample1.xlsx"
bad_file2 = "exampleExcelFiles/UPEI_Ebooks_local_incorrect_sample2.xlsx"
bad_file3 = "exampleExcelFiles/UPEI_Ebooks_local_incorrect_sample3.xlsx"
bad_file4 = "exampleExcelFiles/UPEI_Ebooks_local_incorrect_sample4.xlsx"
bad_file5 = "exampleExcelFiles/UPEI_Ebooks_local_incorrect_sample5.xlsx"
bad_file6 = "exampleExcelFiles/UPEI_Ebooks_local_incorrect_sample6.xlsx"

